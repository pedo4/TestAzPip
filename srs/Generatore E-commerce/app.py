import openai
import os
import requests
import openpyxl
import random
import time
from flask import Flask, render_template, request, send_from_directory

app = Flask(__name__)

num_prod = 0

@app.route('/')
def index():
    # Route per la pagina principale del form E-commerce
    return render_template('form_ECOMM.html')

@app.route('/genera_ecommerce', methods=['POST'])
def genera_ecommerce():
    # Route per la generazione dell'E-commerce
    categorie_selezionate = request.form.getlist('categoria[]') # recupero delle categorie scelte dall'utente
    numeri_prodotti = request.form.getlist('numeroProdotti[]') # recupero del numero dei prodotti per ogni categoria selezionata dall'utente
        
    # Creazione percorso completo di una directory chiamata "static" all'interno dell'applicazione
    templates_dir = os.path.join(app.root_path, 'static')

    for categoria, num_prodotti in zip(categorie_selezionate, numeri_prodotti):
        num_prod = int(num_prodotti)

        # Se 'cucina' rientra tra le categorie scelte dall'utente, richiamo la funzione per la creazione del file html contenente i prodotti di tale categoria
        if categoria == 'cucina':
            description = f"Genera {num_prod} prodotti inserendo una descrizione per ogni prodotto, un prezzo per ogni prodotto ed un nome prodotto. Utilizza lo stesso stile di template_ecomm.html."
            html_text = generate_html_from_description(description, categoria, num_prod)
            file_path = os.path.join(templates_dir, 'cucina.html')
            with open(file_path, "w", encoding='utf-8') as file:
                file.write(html_text)

        # Se 'elettronica' rientra tra le categorie scelte dall'utente, richiamo la funzione per la creazione del file html contenente i prodotti di tale categoria
        elif categoria == 'elettronica':
            description = f"Genera {num_prod} prodotti inserendo una descrizione per ogni prodotto, un prezzo per ogni prodotto ed un nome prodotto. Utilizza lo stesso stile di template_ecomm.html."
            html_text = generate_html_from_description(description, categoria, num_prod)
            file_path = os.path.join(templates_dir, 'elettronica.html')
            with open(file_path, "w", encoding='utf-8') as file:
                file.write(html_text)

        # Se 'abbigliamento' rientra tra le categorie scelte dall'utente, richiamo la funzione per la creazione del file html contenente i prodotti di tale categoria
        elif categoria == 'abbigliamento':
            description = f"Genera {num_prod} prodotti inserendo una descrizione per ogni prodotto, un prezzo per ogni prodotto ed un nome prodotto. Utilizza lo stesso stile di template_ecomm.html."
            html_text = generate_html_from_description(description, categoria, num_prod)
            file_path = os.path.join(templates_dir, 'abbigliamento.html')
            with open(file_path, "w", encoding='utf-8') as file:
                file.write(html_text)
        
        else:
            return "Seleziona almeno una categoria"
    
    # Lettura file 'home_ecommerce.html" che rappresenta la pagina principale dell'e-commerce generato
    with open('templates/home_ecommerce.html', 'r', encoding='utf-8') as file:
        content = file.read()

    # Assegnazione dei link alle pagine delle categorie selezionate dall'utente
    navbar_links = ''
    for categoria in categorie_selezionate:
        if categoria == 'cucina':
            navbar_links += '<a href="{{ url_for(\'static\', filename=\'cucina.html\') }}">Cucina</a>'
        elif categoria == 'elettronica':
            navbar_links += '<a href="{{ url_for(\'static\', filename=\'elettronica.html\') }}">Elettronica</a>'
        elif categoria == 'abbigliamento':
            navbar_links += '<a href="{{ url_for(\'static\', filename=\'abbigliamento.html\') }}">Abbigliamento</a>'

    # Aggiunta dei link alla navbar della home page dell'e-commerce
    modified_content = content.replace('<div class="navbar">', f'<div class="navbar">{navbar_links}')
    
    # Sovrascrttura del file home_ecommerce.html con il contenuto modificato
    with open('templates/home_ecommerce.html', 'w', encoding='utf-8') as file:
        file.write(modified_content)

    return render_template("home_ecommerce.html")


# Route per i file statici
@app.route('/static/<path:filename>')
def static_files(filename):
    return send_from_directory('static', filename)

# Funzione per la generazione delle pagine html
def generate_html_from_description(description, categoria, num_prod):
    # Lettura del contenuto del file template_ecomm.html
    with open("template_ecomm.html", "r", encoding='utf-8') as file:
        html_content = file.read()

    # Caricamento del file Excel a seconda della categoria in questione
    if categoria == 'cucina':
        workbook = openpyxl.load_workbook('cucina.xlsx')
    elif categoria == 'elettronica':
        workbook = openpyxl.load_workbook('elettronica.xlsx')
    elif categoria == 'abbigliamento':
        workbook = openpyxl.load_workbook('abbigliamento.xlsx')

    # Acquisizione foglio di lavoro corretto
    foglio = workbook['Foglio1']  # Sostituzione 'Data' con il nome corretto del foglio di lavoro

    # Acquisizione valori della riga come una lista di stringhe
    valori_riga = [str(cell.value) for cell in foglio[1]]
    
    # Inizializzazione liste per i valori da generare e le URL delle immagini
    valori_da_generare = []
    URL_IMMAGINI = []

    ##########################################################################
    random.seed(time.time())
    a = 0  # Estremo inferiore
    b = 11  # Estremo superiore
    numeri_casuali = []
    ##########################################################################
    
    for j in range(num_prod): #
        n_rand = random.randint(a, b)
        while n_rand in numeri_casuali:
            n_rand = random.randint(a, b)
        
        numeri_casuali.append(n_rand)

        valori_da_generare.append(valori_riga[n_rand])

    #################################################################
    
    for i in range(len(valori_da_generare)):        
        url = 'https://pixabay.com/api/'
        params = {
            'key': '37185814-ef82780621b2221530b9e6698',
            'q': valori_da_generare[i],
            'per_page': 10
        }

        response = requests.get(url, params=params)
        data = response.json()

        # Elaborazione risultati della ricerca delle immagini
        if data['totalHits'] > 0:
            first_image = data['hits'][0]  # Prende solo il primo risultato
            URL_IMMAGINI.append(first_image['webformatURL'])
    

    # Composizione del prompt con la descrizione dell'utente e il contenuto del file HTML
    prompt = f"Alterare il codice HTML che ti invio nel modo seguente: {description}\n\n{html_content}. Usa, inoltre {valori_da_generare} per i prodotti da generare e {URL_IMMAGINI} per le immagini da associare ai prodotti, ricordando che l'immagine da associare al prodotto si trova nello stesso indice/posizione. Traduci tutto in italiano ed emetti il prezzp in €."

    # Impostazione della chiave API
    openai.api_key = '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'

    # Chiamata all'API di completamento di OpenAI,
    response = openai.Completion.create(
        engine='text-davinci-003',
        prompt=prompt,
        max_tokens=2600,
        temperature=1
    )

    html_text = response.choices[0].text
    api_usage = response['usage']
    print('Total token consumed: {0}'.format(api_usage['total_tokens'])) # Stampa per visualizzare il numero di token necessari per soddisfare la richiesta

    return html_text

if __name__ == "__main__":
    app.run()